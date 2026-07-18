"""
Excel photo replacement: swap specific cells' photos in the *original*
uploaded .xlsx, byte-for-byte otherwise.

Unlike a from-scratch rebuild (openpyxl `Workbook()` + `.save()`), this
edits the original zip directly -- every other sheet, style, formula, and
non-photo cell is preserved exactly as uploaded, including the "Place in
Cell" rich-value images openpyxl can't round-trip (see excel_parser.py's
module docstring). The photo cell of a *detected* member (found in the group
photo) is touched two ways: if it already has an image, that image is
swapped in place; if it has no image but its sheet already has a drawing
part (i.e. some other cell on the sheet has a floating picture), a brand-new
picture anchor is added for it. A sheet with no drawing part at all -- no
floating picture anywhere on it -- is left untouched for that row: wiring up
a new drawing part from scratch needs a worksheet-XML splice at an exact
schema-required position that's too easy to get subtly wrong, so that edge
case is skipped rather than risking a corrupt file.
"""

import base64
import io
import posixpath
import re
import xml.etree.ElementTree as ET
import zipfile
from typing import NamedTuple

from openpyxl import load_workbook
from PIL import Image

import excel_parser

_CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"

# Keep in sync with output-template/quiz.js PORTRAIT_*_MARGIN.
PORTRAIT_TOP_MARGIN = 0.20
PORTRAIT_BOTTOM_MARGIN = 0.45
PORTRAIT_SIDE_MARGIN = 0.30

for _prefix, _uri in excel_parser._NS.items():
    # Every real .rels file (package convention) declares this namespace as
    # the unprefixed default (`xmlns="..."`, no `pr:`) -- match that when we
    # write one (the common case: every inserted image adds a relationship),
    # rather than ElementTree's literal "pr" registration. ElementTree's
    # global namespace registry only allows one URI to hold the empty/
    # default prefix at a time, so this is a tradeoff against
    # [Content_Types].xml (see `_ensure_png_content_type`) getting an
    # auto-generated "nsN:" prefix in the rarer case it needs rewriting --
    # cosmetically different from Office's own output, but semantically
    # identical and just as valid XML/OPC, so not worth fighting for both.
    ET.register_namespace("" if _prefix == "pr" else _prefix, _uri)


class PhotoReplacement(NamedTuple):
    sheet_name: str
    row_index: int  # 0-based, matches excel_parser.py's sheetRowIndex
    box: dict  # {"x", "y", "w", "h"}, group-photo pixel space


class NoPhotoColumnError(ValueError):
    """Replacements were requested but no photo column was mapped, so there's
    nothing to swap -- distinct from other errors so a caller can surface an
    honest "nothing to do" message instead of a generic failure (or, worse,
    silently returning the untouched file while claiming success)."""


def replace_photos(
    file_bytes: bytes,
    photo_column: str | None,
    replacements: list[PhotoReplacement],
    group_photo_data_url: str,
    group_photo_width: int,
    group_photo_height: int,
) -> tuple[bytes, int]:
    """Return (new .xlsx bytes, number of cells actually swapped/added).

    The output is identical to `file_bytes` except each entry in
    `replacements` has its photo cell swapped (or, if it had no photo yet,
    given a brand-new one -- see module docstring for the one case that's
    still skipped) for a portrait crop of the group photo at its detection
    box (same crop used on the quiz screen). Everything else -- every other
    cell, sheet, and any row not named here -- is untouched. A row that was
    silently skipped is not counted, so the returned count can be less than
    `len(replacements)`.
    """
    if not replacements:
        return file_bytes, 0
    if not photo_column:
        raise NoPhotoColumnError(
            "No photo column was selected in the column mapping, so there's no cell to "
            "put the replacement photos in."
        )

    group_image = _decode_data_url(group_photo_data_url)
    group_photo_width = group_photo_width or group_image.width
    group_photo_height = group_photo_height or group_image.height

    by_sheet: dict[str, list[PhotoReplacement]] = {}
    for r in replacements:
        by_sheet.setdefault(r.sheet_name, []).append(r)

    media_overrides: dict[str, bytes] = {}
    xml_overrides: dict[str, ET.Element] = {}
    drawing_state: dict[str, dict] = {}
    replaced_count = 0

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)

        for sheet_name, sheet_replacements in by_sheet.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            header = excel_parser._sheet_header(sheet)
            if photo_column not in header:
                continue
            col0 = header.index(photo_column)

            sheet_part = excel_parser._find_sheet_part(zf, sheet_name)
            if sheet_part is None:
                continue
            sheet_root = excel_parser._read_xml(zf, sheet_part)
            if sheet_root is None:
                continue

            row_count = len(excel_parser._read_sheet_rows(sheet, header))
            rich_paths = excel_parser._rich_value_cell_image_paths(zf, sheet_root)
            # Resolved once and threaded through to both the reader below and
            # `_insert_floating_image` so a resize and an insert touching the
            # same sheet in this same call mutate the one shared tree, rather
            # than each parsing their own separate copy of the original and
            # silently discarding whichever the other one already changed.
            sheet_drawing = excel_parser._sheet_drawing(zf, sheet_part, sheet_root)
            floating_records = excel_parser._floating_image_records(
                zf, sheet_part, sheet_root, row_count, sheet_drawing=sheet_drawing
            )
            row_tops = excel_parser._row_tops_emu(sheet_root, row_count + 2)
            col_lefts = excel_parser._col_lefts_emu(sheet_root, col0 + 2)

            for r in sheet_replacements:
                new_bytes = _crop_portrait(
                    group_image, r.box, group_photo_width, group_photo_height
                )
                if new_bytes is None:
                    continue

                rich_path = rich_paths.get((r.row_index + 1, col0))
                if rich_path:
                    media_overrides[rich_path] = new_bytes
                    replaced_count += 1
                    continue

                match = next(
                    (
                        rec
                        for rec in floating_records
                        if rec["col"] == col0 and rec["row_index"] == r.row_index
                    ),
                    None,
                )
                if match is None:
                    if _insert_floating_image(
                        sheet_drawing,
                        zf,
                        col0,
                        r.row_index,
                        row_tops,
                        col_lefts,
                        new_bytes,
                        media_overrides,
                        xml_overrides,
                        drawing_state,
                    ):
                        _ensure_png_content_type(zf, xml_overrides)
                        replaced_count += 1
                    continue  # no drawing part on this sheet -- nothing to add to

                media_overrides[match["media_path"]] = new_bytes
                replaced_count += 1

                drawing_touched = False
                if match["src_rect"] is not None:
                    # A stale crop box computed for the old image would
                    # otherwise be reapplied on top of our new crop.
                    blip_fill_el = match["blip_fill_el"]
                    src_rect_el = blip_fill_el.find("a:srcRect", excel_parser._NS)
                    if src_rect_el is not None:
                        blip_fill_el.remove(src_rect_el)
                        drawing_touched = True
                if match["box"] is not None and _resize_floating_image_box(match, new_bytes):
                    drawing_touched = True
                if drawing_touched:
                    xml_overrides[match["drawing_part"]] = match["drawing_root"]

        return _rewrite_zip(zf, media_overrides, xml_overrides), replaced_count


def _rewrite_zip(
    zf: zipfile.ZipFile,
    media_overrides: dict[str, bytes],
    xml_overrides: dict[str, ET.Element],
) -> bytes:
    out_buf = io.BytesIO()
    written: set[str] = set()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as out_zf:
        for info in zf.infolist():
            if info.filename in media_overrides:
                data = media_overrides[info.filename]
            elif info.filename in xml_overrides:
                data = ET.tostring(
                    xml_overrides[info.filename], encoding="UTF-8", xml_declaration=True
                )
            else:
                data = zf.read(info.filename)
            out_zf.writestr(info, data)
            written.add(info.filename)
        # Brand-new parts (a newly inserted image, or a drawing's .rels file
        # if it didn't have one yet) that aren't in the original zip at all.
        for name, data in media_overrides.items():
            if name not in written:
                out_zf.writestr(name, data)
        for name, el in xml_overrides.items():
            if name not in written:
                out_zf.writestr(name, ET.tostring(el, encoding="UTF-8", xml_declaration=True))
    return out_buf.getvalue()


def _resize_floating_image_box(match: dict, new_bytes: bytes) -> bool:
    """Rewrite a floating picture's anchor geometry so its displayed box
    matches the new image's aspect ratio *and* sits centered within the
    picture's actual cell, instead of leaving the old box (positioned and
    sized for the previous photo, not necessarily centered in its column)
    to stretch the new one out of shape or drift off to one side.

    Keeps the row-height-driven height (`cy`) fixed -- resizing that would
    mean growing/shrinking the row, which is shared by every other cell in
    it -- and derives width from the new image's own aspect ratio, centered
    horizontally in the picture's column. If that width would still be
    wider than the column (an unusually squat crop, or a narrow column),
    both dimensions are shrunk together, still preserving the new aspect
    ratio, so the picture can never spill into the next cell -- it only
    ever gets *smaller* than the row-driven height, never taller. The
    from/to (or ext) anchor and the redundant cached `a:xfrm` box in spPr
    are updated together so they stay consistent with each other.
    """
    box = match["box"]
    new_w, new_h = Image.open(io.BytesIO(new_bytes)).size
    if new_w <= 0 or new_h <= 0:
        return False

    col_lefts = match["col_lefts"]
    col = match["col"]
    cell_left, cell_right = col_lefts[col], col_lefts[col + 1]
    cell_width = cell_right - cell_left

    cy = box["cy"]
    cx = round(cy * new_w / new_h)
    origin_y = box["origin_y"]
    if cx > cell_width:
        cx = cell_width
        shrunk_cy = round(cx * new_h / new_w)
        origin_y += (cy - shrunk_cy) / 2
        cy = shrunk_cy

    origin_x = cell_left + (cell_width - cx) / 2

    from_col, from_col_off = excel_parser._emu_to_col(col_lefts, origin_x)
    _set_child_text(match["from_el"], "xdr:col", from_col)
    _set_child_text(match["from_el"], "xdr:colOff", from_col_off)

    row_tops = match["row_tops"]
    from_row, from_row_off = excel_parser._emu_to_row(row_tops, origin_y)
    _set_child_text(match["from_el"], "xdr:row", from_row)
    _set_child_text(match["from_el"], "xdr:rowOff", from_row_off)

    if match["to_el"] is not None:
        to_col, to_col_off = excel_parser._emu_to_col(col_lefts, origin_x + cx)
        _set_child_text(match["to_el"], "xdr:col", to_col)
        _set_child_text(match["to_el"], "xdr:colOff", to_col_off)
        to_row, to_row_off = excel_parser._emu_to_row(row_tops, origin_y + cy)
        _set_child_text(match["to_el"], "xdr:row", to_row)
        _set_child_text(match["to_el"], "xdr:rowOff", to_row_off)
    elif match["ext_el"] is not None:
        match["ext_el"].set("cx", str(round(cx)))
        match["ext_el"].set("cy", str(round(cy)))

    if match["xfrm_el"] is not None:
        off_el = match["xfrm_el"].find("a:off", excel_parser._NS)
        ext_el = match["xfrm_el"].find("a:ext", excel_parser._NS)
        if off_el is not None:
            off_el.set("x", str(round(origin_x)))
            off_el.set("y", str(round(origin_y)))
        if ext_el is not None:
            ext_el.set("cx", str(round(cx)))
            ext_el.set("cy", str(round(cy)))
    return True


def _set_child_text(parent: ET.Element, tag: str, value: int) -> None:
    el = parent.find(tag, excel_parser._NS)
    if el is not None:
        el.text = str(round(value))


def _insert_floating_image(
    sheet_drawing: tuple[str, ET.Element] | None,
    zf: zipfile.ZipFile,
    col: int,
    row_index: int,
    row_tops: list[int],
    col_lefts: list[int],
    new_bytes: bytes,
    media_overrides: dict[str, bytes],
    xml_overrides: dict[str, ET.Element],
    drawing_state: dict[str, dict],
) -> bool:
    """Add a brand-new floating picture anchored at (col, row_index), sized
    to contain-fit and center within that cell -- only possible when the
    sheet already has a drawing part (see module docstring for why a sheet
    with no drawing part at all is left alone). Returns False, doing
    nothing, if `sheet_drawing` is None (no drawing part to add to).

    `sheet_drawing` must be the *same* (drawing part, drawing root) pair the
    caller is also using for reading existing pictures on this sheet (see
    `_floating_image_records`'s `sheet_drawing` param) -- reusing the one
    shared, already-parsed tree rather than each side re-parsing its own
    copy of the original is what lets a resize and an insert on the same
    sheet, in the same export, both land in the output instead of one
    silently clobbering the other. `drawing_state` similarly carries the
    drawing's `.rels` root across multiple insertions into the same
    drawing within one export, so they don't hand out a duplicate rId.
    """
    if sheet_drawing is None:
        return False
    drawing_part, drawing_root = sheet_drawing

    if drawing_part not in drawing_state:
        rels_path = _rels_path(drawing_part)
        rels_root = excel_parser._read_xml(zf, rels_path)
        if rels_root is None:
            rels_root = ET.Element(f"{{{excel_parser._NS['pr']}}}Relationships")
        drawing_state[drawing_part] = {"rels_path": rels_path, "rels_root": rels_root}
    state = drawing_state[drawing_part]
    rels_root = state["rels_root"]

    new_w, new_h = Image.open(io.BytesIO(new_bytes)).size
    if new_w <= 0 or new_h <= 0:
        return False

    cell_left, cell_right = col_lefts[col], col_lefts[col + 1]
    cell_top, cell_bottom = row_tops[row_index + 1], row_tops[row_index + 2]
    cell_w, cell_h = cell_right - cell_left, cell_bottom - cell_top
    if cell_w <= 0 or cell_h <= 0:
        return False

    if new_w / new_h > cell_w / cell_h:
        cx, cy = cell_w, round(cell_w * new_h / new_w)
    else:
        cy, cx = cell_h, round(cell_h * new_w / new_h)
    origin_x = cell_left + (cell_w - cx) / 2
    origin_y = cell_top + (cell_h - cy) / 2

    media_path = _next_media_path(zf, media_overrides)
    media_overrides[media_path] = new_bytes

    rel_id = _next_rel_id(rels_root)
    ET.SubElement(
        rels_root,
        f"{{{excel_parser._NS['pr']}}}Relationship",
        {
            "Id": rel_id,
            "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image",
            "Target": f"../media/{posixpath.basename(media_path)}",
        },
    )
    xml_overrides[state["rels_path"]] = rels_root

    from_col, from_col_off = excel_parser._emu_to_col(col_lefts, origin_x)
    to_col, to_col_off = excel_parser._emu_to_col(col_lefts, origin_x + cx)
    from_row, from_row_off = excel_parser._emu_to_row(row_tops, origin_y)
    to_row, to_row_off = excel_parser._emu_to_row(row_tops, origin_y + cy)
    anchor = _build_anchor(
        from_col=from_col,
        from_col_off=from_col_off,
        from_row=from_row,
        from_row_off=from_row_off,
        to_col=to_col,
        to_col_off=to_col_off,
        to_row=to_row,
        to_row_off=to_row_off,
        rel_id=rel_id,
        cnvpr_id=_next_cnvpr_id(drawing_root),
        off_x=origin_x,
        off_y=origin_y,
        cx=cx,
        cy=cy,
    )
    drawing_root.append(anchor)
    xml_overrides[drawing_part] = drawing_root
    return True


def _build_anchor(
    *,
    from_col: int,
    from_col_off: int,
    from_row: int,
    from_row_off: int,
    to_col: int,
    to_col_off: int,
    to_row: int,
    to_row_off: int,
    rel_id: str,
    cnvpr_id: int,
    off_x: float,
    off_y: float,
    cx: float,
    cy: float,
) -> ET.Element:
    xdr, a, r = excel_parser._NS["xdr"], excel_parser._NS["a"], excel_parser._NS["r"]

    anchor = ET.Element(f"{{{xdr}}}twoCellAnchor", {"editAs": "oneCell"})
    from_el = ET.SubElement(anchor, f"{{{xdr}}}from")
    ET.SubElement(from_el, f"{{{xdr}}}col").text = str(from_col)
    ET.SubElement(from_el, f"{{{xdr}}}colOff").text = str(from_col_off)
    ET.SubElement(from_el, f"{{{xdr}}}row").text = str(from_row)
    ET.SubElement(from_el, f"{{{xdr}}}rowOff").text = str(from_row_off)
    to_el = ET.SubElement(anchor, f"{{{xdr}}}to")
    ET.SubElement(to_el, f"{{{xdr}}}col").text = str(to_col)
    ET.SubElement(to_el, f"{{{xdr}}}colOff").text = str(to_col_off)
    ET.SubElement(to_el, f"{{{xdr}}}row").text = str(to_row)
    ET.SubElement(to_el, f"{{{xdr}}}rowOff").text = str(to_row_off)

    pic = ET.SubElement(anchor, f"{{{xdr}}}pic")
    nv_pic_pr = ET.SubElement(pic, f"{{{xdr}}}nvPicPr")
    ET.SubElement(
        nv_pic_pr, f"{{{xdr}}}cNvPr", {"id": str(cnvpr_id), "name": f"Picture {cnvpr_id}"}
    )
    cnv_pic_pr = ET.SubElement(nv_pic_pr, f"{{{xdr}}}cNvPicPr")
    ET.SubElement(cnv_pic_pr, f"{{{a}}}picLocks", {"noChangeAspect": "1"})

    blip_fill = ET.SubElement(pic, f"{{{xdr}}}blipFill")
    ET.SubElement(blip_fill, f"{{{a}}}blip", {f"{{{r}}}embed": rel_id})
    stretch = ET.SubElement(blip_fill, f"{{{a}}}stretch")
    ET.SubElement(stretch, f"{{{a}}}fillRect")

    sp_pr = ET.SubElement(pic, f"{{{xdr}}}spPr")
    xfrm = ET.SubElement(sp_pr, f"{{{a}}}xfrm")
    ET.SubElement(xfrm, f"{{{a}}}off", {"x": str(round(off_x)), "y": str(round(off_y))})
    ET.SubElement(xfrm, f"{{{a}}}ext", {"cx": str(round(cx)), "cy": str(round(cy))})
    prst_geom = ET.SubElement(sp_pr, f"{{{a}}}prstGeom", {"prst": "rect"})
    ET.SubElement(prst_geom, f"{{{a}}}avLst")

    ET.SubElement(anchor, f"{{{xdr}}}clientData")
    return anchor


def _rels_path(part_name: str) -> str:
    directory, filename = posixpath.split(part_name)
    return posixpath.join(directory, "_rels", filename + ".rels")


def _next_media_path(zf: zipfile.ZipFile, media_overrides: dict[str, bytes]) -> str:
    existing = set(zf.namelist()) | set(media_overrides.keys())
    max_idx = 0
    for name in existing:
        m = re.match(r"xl/media/image(\d+)\.", name)
        if m:
            max_idx = max(max_idx, int(m.group(1)))
    return f"xl/media/image{max_idx + 1}.png"


def _next_rel_id(rels_root: ET.Element) -> str:
    max_idx = 0
    for rel in rels_root.findall("pr:Relationship", excel_parser._NS):
        m = re.match(r"rId(\d+)", rel.get("Id") or "")
        if m:
            max_idx = max(max_idx, int(m.group(1)))
    return f"rId{max_idx + 1}"


def _next_cnvpr_id(drawing_root: ET.Element) -> int:
    max_idx = 0
    for cnvpr in drawing_root.findall(".//xdr:cNvPr", excel_parser._NS):
        try:
            max_idx = max(max_idx, int(cnvpr.get("id")))
        except (TypeError, ValueError):
            pass
    return max_idx + 1


def _ensure_png_content_type(zf: zipfile.ZipFile, xml_overrides: dict[str, ET.Element]) -> None:
    """The new picture is always saved as PNG (see `_crop_portrait`); a
    workbook whose existing images happen to all be some other format
    (e.g. only .jpeg) won't have declared a Default content type for
    "png" yet, and a media part covered by neither a Default nor an
    Override extension is an invalid OPC package. No-ops if already
    declared, including by an earlier insertion in this same export.
    """
    root = xml_overrides.get("[Content_Types].xml") or excel_parser._read_xml(
        zf, "[Content_Types].xml"
    )
    if root is None:
        return
    has_png = any(
        d.get("Extension", "").lower() == "png" for d in root.findall(f"{{{_CT_NS}}}Default")
    )
    if not has_png:
        ET.SubElement(root, f"{{{_CT_NS}}}Default", {"Extension": "png", "ContentType": "image/png"})
        xml_overrides["[Content_Types].xml"] = root


def _crop_portrait(
    group_image: Image.Image, box: dict, image_width: int, image_height: int
) -> bytes | None:
    x, y, w, h = box["x"], box["y"], box["w"], box["h"]
    top = max(0, y - h * PORTRAIT_TOP_MARGIN)
    bottom = min(image_height, y + h + h * PORTRAIT_BOTTOM_MARGIN)
    left = max(0, x - w * PORTRAIT_SIDE_MARGIN)
    right = min(image_width, x + w + w * PORTRAIT_SIDE_MARGIN)
    if right <= left or bottom <= top:
        return None
    box_px = (round(left), round(top), round(right), round(bottom))
    cropped = group_image.crop(box_px)
    buf = io.BytesIO()
    cropped.convert("RGB").save(buf, format="PNG")
    return buf.getvalue()


def _decode_data_url(data_url: str) -> Image.Image:
    _, encoded = data_url.split(",", 1)
    return Image.open(io.BytesIO(base64.b64decode(encoded))).convert("RGB")
