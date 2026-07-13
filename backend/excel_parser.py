"""
Excel import (openpyxl for cell values, raw OOXML parsing for images).

Responsibilities:
- Read uploaded .xlsx, return column headers to frontend for mapping.
- Given chosen columns (name / division / photo / answer-text columns),
  build a list of member records.
- Assign a hidden internal ID to every row (not just duplicates).
- Extract images bound to cells and match each to its row -> return match
  list for the frontend's preview/confirm grid.

A photo can reach a cell two different ways, and a single sheet may mix
both (e.g. most rows use one method, one row was fixed up with the other):
  1. Excel's "Place in Cell" picture feature, where the image genuinely IS
     the cell value. openpyxl has no support for this at all -- the cell
     just reads back as the literal string "#VALUE!". This is reconstructed
     here by walking the OOXML rich-value chain: a cell's `vm` attribute ->
     xl/metadata.xml valueMetadata -> xl/richData/rdrichvalue.xml -> a
     richValueRel index -> xl/richData/richValueRel.xml -> a relationship id
     -> xl/media/imageN.*. This is an exact cell match and takes priority.
  2. A classic floating picture anchored near a cell
     (xl/drawings/drawingN.xml), read directly rather than via openpyxl's
     private `sheet._images`. A floating picture doesn't necessarily line up
     with row boundaries, so it's matched to whichever row its vertical span
     overlaps the most (by % of the picture's own height), and is only used
     as a fallback for rows that have no "Place in Cell" photo. If it was
     cropped in Excel, the embedded media is still the full original image
     (Excel just stores a <a:srcRect> crop box) -- that crop is re-applied
     here so the extracted photo matches what's visible in the sheet.
"""

import base64
import io
import posixpath
import re
import uuid
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from PIL import Image

_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "rd": "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata",
    "rvrel": "http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel",
}
_R_ID = f"{{{_NS['r']}}}id"
_CELL_REF_RE = re.compile(r"([A-Za-z]+)(\d+)")
_EXT_MIME = {
    "png": "png", "jpg": "jpeg", "jpeg": "jpeg", "gif": "gif",
    "bmp": "bmp", "tif": "tiff", "tiff": "tiff", "webp": "webp",
}


def parse_excel(file_bytes: bytes) -> dict:
    """Read the first sheet of an .xlsx file.

    Returns:
        {
          "columns": ["Name", "Division", ...],
          "rows": [{"id": "<uuid>", "cells": {"Name": "Alice", ...}}, ...],
          "photoMatches": {"<row id>": "data:image/png;base64,..." | None, ...},
        }
    """
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    header = next(rows_iter, None)
    columns = [str(cell) if cell is not None else "" for cell in (header or [])]

    rows = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        cells = {
            columns[i]: values[i] if i < len(values) else None
            for i in range(len(columns))
        }
        rows.append({"id": str(uuid.uuid4()), "cells": cells})

    photo_matches = _match_images_to_rows(file_bytes, sheet.title, len(rows))
    return {
        "columns": columns,
        "rows": rows,
        "photoMatches": {
            rows[i]["id"]: photo_matches.get(i) for i in range(len(rows))
        },
    }


def _match_images_to_rows(
    file_bytes: bytes, sheet_title: str, row_count: int
) -> dict[int, str]:
    """Match every image bound to a cell to its data row. Header occupies
    sheet row 0; data row `i` sits at sheet row `i + 1`. `photo_col` is
    inferred as the most common image column, since a sheet normally has
    exactly one photo column. "Place in Cell" images are an exact cell
    match and always win; floating images (matched by row overlap %) only
    fill rows that got no "Place in Cell" photo.
    """
    if row_count == 0:
        return {}

    with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
        sheet_part = _find_sheet_part(zf, sheet_title)
        if sheet_part is None:
            return {}
        sheet_root = _read_xml(zf, sheet_part)
        if sheet_root is None:
            return {}

        rich_value_images = _rich_value_cell_images(zf, sheet_root)
        floating_images = _floating_images(zf, sheet_part, sheet_root, row_count)

    photo_col_votes = [col for _, col in rich_value_images] or [
        col for col, _, _ in floating_images
    ]
    if not photo_col_votes:
        return {}
    photo_col = Counter(photo_col_votes).most_common(1)[0][0]

    matches: dict[int, str] = {}
    for (row, col), data_url in rich_value_images.items():
        row_index = row - 1
        if col == photo_col and 0 <= row_index < row_count:
            matches[row_index] = data_url

    for col, row_index, data_url in floating_images:
        if col == photo_col and row_index not in matches:
            matches[row_index] = data_url

    return matches


def _rich_value_cell_images(
    zf: zipfile.ZipFile, sheet_root: ET.Element
) -> dict[tuple[int, int], str]:
    """"Place in Cell" pictures: the cell's own value is the image."""
    vm_cells = [
        (rc, int(c_el.get("vm")))
        for c_el in sheet_root.findall(".//main:c", _NS)
        if c_el.get("vm") and (rc := _cell_ref_to_rc(c_el.get("r", "")))
    ]
    if not vm_cells:
        return {}

    vm_to_rv = _rich_value_index_by_vm(zf)
    rv_paths = _rich_value_image_paths(zf)
    if not rv_paths:
        return {}

    images: dict[tuple[int, int], str] = {}
    for rc, vm in vm_cells:
        rv_index = vm_to_rv.get(vm)
        path = rv_paths.get(rv_index) if rv_index is not None else None
        if path and path in zf.namelist():
            images[rc] = _data_url(zf.read(path), path)
    return images


def _rich_value_index_by_vm(zf: zipfile.ZipFile) -> dict[int, int]:
    """{vm attribute value (1-based) -> rv index in rdrichvalue.xml}."""
    metadata_root = _read_xml(zf, "xl/metadata.xml")
    if metadata_root is None:
        return {}

    rich_type_index = None
    for i, type_el in enumerate(
        metadata_root.findall("main:metadataTypes/main:metadataType", _NS), start=1
    ):
        if type_el.get("name") == "XLRICHVALUE":
            rich_type_index = i
            break
    if rich_type_index is None:
        return {}

    result: dict[int, int] = {}
    bk_list = metadata_root.findall("main:valueMetadata/main:bk", _NS)
    for vm_index, bk_el in enumerate(bk_list, start=1):
        for rc_el in bk_el.findall("main:rc", _NS):
            if rc_el.get("t") == str(rich_type_index):
                try:
                    result[vm_index] = int(rc_el.get("v"))
                except (TypeError, ValueError):
                    pass
                break
    return result


def _rich_value_image_paths(zf: zipfile.ZipFile) -> dict[int, str]:
    """{rv index in rdrichvalue.xml -> zip path of the referenced image}."""
    rv_root = _read_xml(zf, "xl/richData/rdrichvalue.xml")
    struct_root = _read_xml(zf, "xl/richData/rdrichvaluestructure.xml")
    rel_root = _read_xml(zf, "xl/richData/richValueRel.xml")
    if rv_root is None or struct_root is None or rel_root is None:
        return {}

    # Each rv's <v> children are positional per its structure; find where
    # the image-identifier field sits within its declared structure.
    image_key_position: dict[int, int] = {}
    for struct_index, struct_el in enumerate(struct_root.findall("rd:s", _NS)):
        for pos, key_el in enumerate(struct_el.findall("rd:k", _NS)):
            if "imageidentifier" in (key_el.get("n") or "").lower():
                image_key_position[struct_index] = pos
                break

    rv_to_rel_index: dict[int, int] = {}
    for rv_index, rv_el in enumerate(rv_root.findall("rd:rv", _NS)):
        key_pos = image_key_position.get(int(rv_el.get("s", "-1")))
        if key_pos is None:
            continue
        values = rv_el.findall("rd:v", _NS)
        if key_pos < len(values):
            try:
                rv_to_rel_index[rv_index] = int(values[key_pos].text)
            except (TypeError, ValueError):
                continue

    rel_ids = [
        rel_el.get(_R_ID) for rel_el in rel_root.findall("rvrel:rel", _NS)
    ]
    rel_targets = _read_rels(zf, "xl/richData/richValueRel.xml")

    paths: dict[int, str] = {}
    for rv_index, rel_index in rv_to_rel_index.items():
        if 0 <= rel_index < len(rel_ids):
            path = rel_targets.get(rel_ids[rel_index])
            if path:
                paths[rv_index] = path
    return paths


def _floating_images(
    zf: zipfile.ZipFile, sheet_part: str, sheet_root: ET.Element, row_count: int
) -> list[tuple[int, int, str]]:
    """Classic pictures anchored near a cell, via the sheet's drawing.

    A picture's anchor doesn't have to land exactly on a row boundary (drag
    handles move in pixels, not rows), so each image is assigned to whichever
    data row its vertical span overlaps the most, as a percentage of the
    image's own height. Returns (col, row_index, data-url) triples.
    """
    drawing_el = sheet_root.find("main:drawing", _NS)
    if drawing_el is None:
        return []
    drawing_part = _read_rels(zf, sheet_part).get(drawing_el.get(_R_ID))
    drawing_root = _read_xml(zf, drawing_part) if drawing_part else None
    if drawing_root is None:
        return []
    drawing_rels = _read_rels(zf, drawing_part)

    anchors = drawing_root.findall("xdr:twoCellAnchor", _NS) + drawing_root.findall(
        "xdr:oneCellAnchor", _NS
    )
    parsed = []
    max_row_seen = row_count + 1
    for anchor_el in anchors:
        blip_fill_el = anchor_el.find(".//xdr:blipFill", _NS)
        blip_el = blip_fill_el.find("a:blip", _NS) if blip_fill_el is not None else None
        from_el = anchor_el.find("xdr:from", _NS)
        if blip_el is None or from_el is None:
            continue
        embed_id = blip_el.get(f"{{{_NS['r']}}}embed")
        col = _int_or_none(from_el.findtext("xdr:col", namespaces=_NS))
        from_row = _int_or_none(from_el.findtext("xdr:row", namespaces=_NS))
        from_off = _int_or_none(from_el.findtext("xdr:rowOff", namespaces=_NS)) or 0
        if not embed_id or col is None or from_row is None:
            continue
        path = drawing_rels.get(embed_id)
        if not path or path not in zf.namelist():
            continue
        src_rect = _parse_src_rect(blip_fill_el.find("a:srcRect", _NS))

        to_el = anchor_el.find("xdr:to", _NS)
        to_row = to_off = ext_cy = None
        if to_el is not None:
            to_row = _int_or_none(to_el.findtext("xdr:row", namespaces=_NS))
            to_off = _int_or_none(to_el.findtext("xdr:rowOff", namespaces=_NS)) or 0
        else:
            ext_el = anchor_el.find("xdr:ext", _NS)
            if ext_el is not None and ext_el.get("cy"):
                ext_cy = int(ext_el.get("cy"))

        max_row_seen = max(max_row_seen, from_row, to_row or 0)
        parsed.append((col, from_row, from_off, to_row, to_off, ext_cy, path, src_rect))

    if not parsed:
        return []

    row_tops = _row_tops_emu(sheet_root, max_row_seen + 2)

    results = []
    for col, from_row, from_off, to_row, to_off, ext_cy, path, src_rect in parsed:
        img_top = row_tops[from_row] + from_off
        img_bottom = (
            row_tops[to_row] + to_off if to_row is not None else img_top + (ext_cy or 0)
        )
        row_index = _best_overlapping_row(img_top, img_bottom, row_tops, row_count)
        if row_index is not None:
            data = _crop_to_src_rect(zf.read(path), src_rect)
            results.append((col, row_index, _data_url(data, path)))
    return results


def _parse_src_rect(src_rect_el: ET.Element | None) -> tuple[float, float, float, float] | None:
    """Excel's crop tool doesn't touch the embedded media file -- it stores
    how much of each edge is trimmed as a <a:srcRect> on the picture's
    blipFill, in thousandths of a percent (l/t/r/b, 100000 = 100%). Returns
    (left, top, right, bottom) as 0..1 fractions, or None if uncropped.
    """
    if src_rect_el is None:
        return None
    fractions = tuple(
        max(0, int(src_rect_el.get(edge, "0"))) / 100000 for edge in ("l", "t", "r", "b")
    )
    return fractions if any(fractions) else None


def _crop_to_src_rect(
    data: bytes, src_rect: tuple[float, float, float, float] | None
) -> bytes:
    if src_rect is None:
        return data
    left, top, right, bottom = src_rect
    try:
        image = Image.open(io.BytesIO(data))
        width, height = image.size
        box = (
            round(width * left),
            round(height * top),
            round(width * (1 - right)),
            round(height * (1 - bottom)),
        )
        if box[0] >= box[2] or box[1] >= box[3]:
            return data
        cropped = image.crop(box)
        buf = io.BytesIO()
        cropped.save(buf, format=image.format or "PNG")
        return buf.getvalue()
    except Exception:
        return data


def _row_tops_emu(sheet_root: ET.Element, num_rows: int) -> list[int]:
    """EMU offset of the top of each 0-indexed row, up to `num_rows`
    (tops[i] is the top of row i; tops[-1] is the bottom of the last row).
    """
    default_pt = 15.0
    fmt_el = sheet_root.find("main:sheetFormatPr", _NS)
    if fmt_el is not None and fmt_el.get("defaultRowHeight"):
        default_pt = float(fmt_el.get("defaultRowHeight"))

    heights_pt = {}
    for row_el in sheet_root.findall("main:sheetData/main:row", _NS):
        if row_el.get("ht"):
            heights_pt[int(row_el.get("r")) - 1] = float(row_el.get("ht"))

    tops = [0.0]
    for i in range(num_rows):
        tops.append(tops[-1] + heights_pt.get(i, default_pt))
    return [int(round(pt * 12700)) for pt in tops]


def _best_overlapping_row(
    img_top: int, img_bottom: int, row_tops: list[int], row_count: int
) -> int | None:
    """Data row index whose vertical span overlaps the image the most, as a
    percentage of the image's own height. `None` if the image doesn't
    overlap any data row (e.g. it sits entirely over the header).
    """
    img_height = img_bottom - img_top
    if img_height <= 0:
        return None
    best_index, best_pct = None, 0.0
    for row_index in range(row_count):
        xml_row = row_index + 1  # data row i sits at sheet row i + 1
        row_top, row_bottom = row_tops[xml_row], row_tops[xml_row + 1]
        overlap = min(img_bottom, row_bottom) - max(img_top, row_top)
        pct = overlap / img_height
        if pct > best_pct:
            best_pct, best_index = pct, row_index
    return best_index


def _int_or_none(text: str | None) -> int | None:
    if text is None:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _find_sheet_part(zf: zipfile.ZipFile, sheet_title: str) -> str | None:
    workbook_root = _read_xml(zf, "xl/workbook.xml")
    if workbook_root is None:
        return None
    rel_id = next(
        (
            sheet_el.get(_R_ID)
            for sheet_el in workbook_root.findall("main:sheets/main:sheet", _NS)
            if sheet_el.get("name") == sheet_title
        ),
        None,
    )
    if rel_id is None:
        return None
    return _read_rels(zf, "xl/workbook.xml").get(rel_id)


def _read_xml(zf: zipfile.ZipFile, name: str | None) -> ET.Element | None:
    if not name or name not in zf.namelist():
        return None
    return ET.fromstring(zf.read(name))


def _read_rels(zf: zipfile.ZipFile, part_name: str) -> dict[str, str]:
    """{relationship id -> zip path} for the .rels file of `part_name`."""
    directory, filename = posixpath.split(part_name)
    rels_root = _read_xml(zf, posixpath.join(directory, "_rels", filename + ".rels"))
    if rels_root is None:
        return {}
    rels = {}
    for rel_el in rels_root.findall("pr:Relationship", _NS):
        target, rel_id = rel_el.get("Target"), rel_el.get("Id")
        if target and rel_id:
            rels[rel_id] = posixpath.normpath(posixpath.join(directory, target))
    return rels


def _cell_ref_to_rc(ref: str) -> tuple[int, int] | None:
    match = _CELL_REF_RE.match(ref)
    if not match:
        return None
    col_letters, row_digits = match.groups()
    return int(row_digits) - 1, column_index_from_string(col_letters) - 1


def _data_url(data: bytes, path: str) -> str:
    ext = path.rsplit(".", 1)[-1].lower() if "." in path else "png"
    mime = _EXT_MIME.get(ext, "png")
    encoded = base64.b64encode(data).decode("ascii")
    return f"data:image/{mime};base64,{encoded}"
